"""
Waitlist Router for JewelTech
Handles waitlist signup, Google Sheets export, and Excel download
"""
from fastapi import APIRouter, HTTPException, status, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from backend.models.mongodb import UserModel, WaitlistModel, TrialUsageModel
from backend.utils.auth import (
    get_current_user, get_admin_user, EmailService
)
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()


class JoinWaitlistRequest(BaseModel):
    """Request to join waitlist (user must be authenticated)"""
    pass


class WaitlistResponse(BaseModel):
    message: str
    position: int
    total_count: int
    in_waitlist: bool


@router.post("/join", response_model=WaitlistResponse)
async def join_waitlist(
    current_user: dict = Depends(get_current_user)
):
    """
    Add current user to waitlist
    (Does not require email verification - users can join waitlist before verifying)
    """
    user_id = current_user["_id"]
    email = current_user["email"]

    # Check if already in waitlist
    if WaitlistModel.is_in_waitlist(user_id):
        total_count = WaitlistModel.get_waitlist_count()
        return {
            "message": "You're already on the waitlist!",
            "position": 0,  # Could calculate actual position
            "total_count": total_count,
            "in_waitlist": True
        }

    # Get trial usage summary
    features = ["ai_designer", "virtual_tryon", "qc_inspector", "3d_generation"]
    trial_usage_summary = {}

    for feature in features:
        usage_info = TrialUsageModel.check_trial_limit(user_id, feature)
        trial_usage_summary[feature] = {
            "used": usage_info.get("used", 0),
            "limit": usage_info.get("limit", 3)
        }

    # Add to waitlist
    waitlist_data = {
        "full_name": current_user.get("full_name"),
        "phone": current_user.get("phone"),
        "username": current_user.get("username"),
        "trial_usage_summary": trial_usage_summary
    }

    entry = WaitlistModel.add_to_waitlist(user_id, email, waitlist_data)

    # Get position
    total_count = WaitlistModel.get_waitlist_count()

    # Send confirmation email
    EmailService.send_waitlist_confirmation(
        email,
        current_user["username"],
        total_count
    )

    # Try to sync to Google Sheets (non-blocking, don't fail if it errors)
    try:
        await sync_to_google_sheets()
    except Exception as e:
        print(f"Google Sheets sync error: {e}")

    return {
        "message": "Successfully joined the waitlist!",
        "position": total_count,
        "total_count": total_count,
        "in_waitlist": True
    }


@router.get("/status")
async def get_waitlist_status(
    current_user: dict = Depends(get_current_user)
):
    """
    Check if current user is on waitlist
    (Does not require email verification - users can check status before verifying)
    """
    user_id = current_user["_id"]
    in_waitlist = WaitlistModel.is_in_waitlist(user_id)
    total_count = WaitlistModel.get_waitlist_count()

    return {
        "in_waitlist": in_waitlist,
        "total_count": total_count,
        "username": current_user["username"]
    }


@router.get("/export-excel")
async def export_waitlist_to_excel(
    admin_user: dict = Depends(get_admin_user)
):
    """
    Export waitlist to Excel file (Admin only)
    """
    # Get all waitlist entries
    entries = WaitlistModel.get_waitlist(limit=10000)

    if not entries:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No waitlist entries found"
        )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "JewelTech Waitlist"

    # Header styling
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Position", "Username", "Email", "Full Name", "Phone",
        "Joined Date", "Status", "AI Designer Used", "Try-On Used",
        "QC Inspector Used", "3D Gen Used"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for idx, entry in enumerate(entries, 2):
        trial_summary = entry.get("trial_usage_summary", {})

        ws.cell(row=idx, column=1, value=idx - 1)  # Position
        ws.cell(row=idx, column=2, value=entry.get("username", ""))
        ws.cell(row=idx, column=3, value=entry.get("email", ""))
        ws.cell(row=idx, column=4, value=entry.get("full_name", ""))
        ws.cell(row=idx, column=5, value=entry.get("phone", ""))
        ws.cell(row=idx, column=6, value=entry.get("joined_at", "").strftime("%Y-%m-%d %H:%M") if entry.get("joined_at") else "")
        ws.cell(row=idx, column=7, value=entry.get("status", "pending"))

        # Trial usage
        ws.cell(row=idx, column=8, value=f"{trial_summary.get('ai_designer', {}).get('used', 0)}/3")
        ws.cell(row=idx, column=9, value=f"{trial_summary.get('virtual_tryon', {}).get('used', 0)}/3")
        ws.cell(row=idx, column=10, value=f"{trial_summary.get('qc_inspector', {}).get('used', 0)}/3")
        ws.cell(row=idx, column=11, value=f"{trial_summary.get('3d_generation', {}).get('used', 0)}/3")

    # Adjust column widths
    column_widths = [10, 15, 30, 20, 15, 20, 12, 15, 12, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"jeweltech_waitlist_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


async def sync_to_google_sheets():
    """
    Sync waitlist to Google Sheets (requires service account credentials)
    """
    try:
        import gspread
        from oauth2client.service_account import ServiceAccountCredentials

        # Check if credentials file exists
        creds_file = os.getenv("GOOGLE_SHEETS_CREDENTIALS_FILE", "google_credentials.json")
        if not os.path.exists(creds_file):
            print("Google Sheets credentials not found. Skipping sync.")
            return

        # Authenticate
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_file, scope)
        client = gspread.authorize(creds)

        # Open spreadsheet (or create if doesn't exist)
        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        if not sheet_id:
            print("GOOGLE_SHEETS_ID not set. Skipping sync.")
            return

        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.sheet1

        # Get waitlist entries
        entries = WaitlistModel.get_waitlist(limit=10000)

        # Clear existing data (except header)
        worksheet.clear()

        # Headers
        headers = [
            "Position", "Username", "Email", "Full Name", "Phone",
            "Joined Date", "Status", "AI Designer", "Try-On",
            "QC Inspector", "3D Gen"
        ]
        worksheet.append_row(headers)

        # Data rows
        for idx, entry in enumerate(entries, 1):
            trial_summary = entry.get("trial_usage_summary", {})
            row = [
                idx,
                entry.get("username", ""),
                entry.get("email", ""),
                entry.get("full_name", ""),
                entry.get("phone", ""),
                entry.get("joined_at", "").strftime("%Y-%m-%d %H:%M") if entry.get("joined_at") else "",
                entry.get("status", "pending"),
                f"{trial_summary.get('ai_designer', {}).get('used', 0)}/3",
                f"{trial_summary.get('virtual_tryon', {}).get('used', 0)}/3",
                f"{trial_summary.get('qc_inspector', {}).get('used', 0)}/3",
                f"{trial_summary.get('3d_generation', {}).get('used', 0)}/3"
            ]
            worksheet.append_row(row)

        print("Successfully synced to Google Sheets!")

    except Exception as e:
        print(f"Error syncing to Google Sheets: {e}")
        raise


@router.get("/list")
async def get_waitlist_entries(
    admin_user: dict = Depends(get_admin_user),
    status: Optional[str] = None,
    limit: int = 100
):
    """
    Get waitlist entries (Admin only)
    """
    entries = WaitlistModel.get_waitlist(status=status, limit=limit)

    return {
        "total": len(entries),
        "entries": entries
    }


@router.post("/sync-google-sheets")
async def trigger_google_sheets_sync(
    admin_user: dict = Depends(get_admin_user)
):
    """
    Manually trigger Google Sheets sync (Admin only)
    """
    try:
        await sync_to_google_sheets()
        return {
            "message": "Successfully synced to Google Sheets"
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error syncing to Google Sheets: {str(e)}"
        )
